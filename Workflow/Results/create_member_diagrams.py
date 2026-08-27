"""
Erstellt drei Nachweis-Diagramme (einzeln und als kombinierte Uebersicht) aus einer
rohen "Members_Comparison"-Excel-Datei (Spalte 'key' + 'level', je eine Spalte pro Member) -
demselben Rohformat, wie es struct_analysis.py fuer Rippenplatten (rc_rib) exportiert.

Diagramme:
1. Biegewiderstand mu_max (MRd) vs. Bemessungsmoment mEd_p
2. Querkraftwiderstand vu_p (VRd) vs. Bemessungsquerkraft vEd
3. Rissmoment mr_p vs. mkd_p (positive Biegung)
4. Rissmoment mr_n vs. mkd_n (negative Biegung)
5. Durchbiegung w_install_adm (zulässig) vs. w_install_ger (gerissen)
6. Durchbiegung w_use_adm (zulässig) vs. w_use_ger (gerissen)
7. Durchbiegung w_app_adm (zulässig) vs. w_app_ger (gerissen)

X-Achse je Subplot: Globaler Member-Index, jeder Member nur einmal.
Kategorisierung: Spannweite l_tot (automatisch aus den Daten ermittelt), je ein Subplot.
Hinterlegung (hellrot) markiert den jeweils kritischen Bereich (z.B. Bemessungswert > Widerstand).

Die Excel-Datei wird unten in EXCEL_DATEI definiert. Einfach den Dateinamen anpassen und
das Skript starten (kein Kommandozeilen-Argument noetig). Optional kann beim Aufruf trotzdem
ein Pfad als erstes Argument uebergeben werden - dieser hat dann Vorrang vor EXCEL_DATEI:
    py create_member_diagrams.py [<pfad_zur_excel_datei>] [sheet_name]

Die Diagramme werden im selben Ordner wie die Eingabedatei gespeichert, benannt nach dem
Dateinamen der Excel-Datei (ohne Endung).
"""
import os
import sys

import matplotlib as mpl
import matplotlib.pyplot as plt
import openpyxl
import pandas as pd
from matplotlib.patches import Patch

# Zu analysierende Excel-Datei (liegt im selben Ordner wie dieses Skript) - hier anpassen:
EXCEL_DATEI = "Members_rc_rec_simple_massiv.xlsx"

DEFAULT_SHEET_NAME = "Members_Comparison"

# (Spaltenname im Diagramm, key, level) - level wird benoetigt, da Keys mehrfach vorkommen koennen
BENOETIGTE_FELDER = [
    ("l_tot", "l_tot", 1),
    ("w_install_adm", "w_install_adm", 0),
    ("w_install_ger", "w_install_ger", 0),
    ("mu_max", "mu_max", 1),
    ("mEd_p", "mEd_p", 0),
    ("vu_p", "vu_p", 1),
    ("vEd", "vEd", 0),
    ("w_use_adm", "w_use_adm", 0),
    ("w_use_ger", "w_use_ger", 0),
    ("w_app_adm", "w_app_adm", 0),
    ("w_app_ger", "w_app_ger", 0),
    ("mr_p", "mr_p", 1),
    ("mr_n", "mr_n", 1),
    ("mkd_p", "mkd_p", 0),
    ("mkd_n", "mkd_n", 0),
]

mpl.rcParams['font.family'] = 'serif'
mpl.rcParams['font.serif'] = ['Times New Roman']
mpl.rcParams['font.size'] = 14
mpl.rcParams['axes.labelsize'] = 14
mpl.rcParams['axes.titlesize'] = 14
mpl.rcParams['xtick.labelsize'] = 14
mpl.rcParams['ytick.labelsize'] = 14
mpl.rcParams['legend.fontsize'] = 14

FARBE_1 = "#0072B2"           # Widerstand / Zulässige Durchbiegung
FARBE_2 = "#D55E00"           # Bemessungswert / gerissen
FARBE_HINTERLEGUNG = "#f4a9a9"  # Hinterlegung: kritischer Bereich


# ==============================================================================
# DATENAUFBEREITUNG
# ==============================================================================
def lade_daten(excel_file, sheet_name):
    """Liest die benoetigten Zeilen aus der rohen Members_Comparison-Excel und
    transponiert sie zu einer Tabelle mit einer Zeile je Member."""
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    ws = wb[sheet_name]

    zeilen_index = {}
    for r in range(1, ws.max_row + 1):
        key = ws.cell(row=r, column=1).value
        level = ws.cell(row=r, column=2).value
        for spalte, gesuchter_key, gesuchtes_level in BENOETIGTE_FELDER:
            if spalte not in zeilen_index and key == gesuchter_key and level == gesuchtes_level:
                zeilen_index[spalte] = r

    fehlend = [spalte for spalte, _, _ in BENOETIGTE_FELDER if spalte not in zeilen_index]
    if fehlend:
        raise ValueError(f"Folgende benoetigten Felder wurden in '{excel_file}' nicht gefunden: {fehlend}")

    member_ids = [ws.cell(row=1, column=c).value for c in range(3, ws.max_column + 1)]
    daten = {"Member_ID": member_ids}
    for spalte, r in zeilen_index.items():
        daten[spalte] = [ws.cell(row=r, column=c).value for c in range(3, ws.max_column + 1)]

    df = pd.DataFrame(daten)
    for spalte, _, _ in BENOETIGTE_FELDER:
        df[spalte] = pd.to_numeric(df[spalte], errors="coerce")
    df = df.dropna(subset=[spalte for spalte, _, _ in BENOETIGTE_FELDER])

    # Einheiten: Nm -> kNm, N -> kN fuer bessere Lesbarkeit
    df["mu_max"] = df["mu_max"] / 1000
    df["mEd_p"] = df["mEd_p"] / 1000
    df["vu_p"] = df["vu_p"] / 1000
    df["vEd"] = df["vEd"] / 1000
    df["mr_p"] = df["mr_p"] / 1000
    df["mr_n"] = df["mr_n"] / 1000
    df["mkd_p"] = df["mkd_p"] / 1000
    df["mkd_n"] = df["mkd_n"] / 1000

    # Globaler, eindeutiger Index je Member, abgeleitet aus Member_ID
    df["member_index"] = df["Member_ID"].str.extract(r"(\d+)").astype(int)

    df["w_install_ger_groesser_als_w_install_adm"] = df["w_install_ger"] > df["w_install_adm"]
    df["med_groesser_als_mrd"] = df["mEd_p"] > df["mu_max"]
    df["ved_groesser_als_vrd"] = df["vEd"] > df["vu_p"]
    df["w_use_ger_groesser_als_w_use_adm"] = df["w_use_ger"] > df["w_use_adm"]
    df["w_app_ger_groesser_als_w_app_adm"] = df["w_app_ger"] > df["w_app_adm"]
    df["mkd_p_groesser_als_mr_p"] = df["mkd_p"] > df["mr_p"]
    df["mkd_n_kleiner_als_mr_n"] = df["mkd_n"] < df["mr_n"]

    return df


def shade_condition_regions(ax, x_vals, condition, color, alpha=0.6):
    """Hinterlegt zusammenhaengende x-Bereiche, in denen 'condition' True ist."""
    x_vals = list(x_vals)
    condition = list(condition)
    n = len(x_vals)
    i = 0
    while i < n:
        if condition[i]:
            j = i
            while j + 1 < n and condition[j + 1]:
                j += 1
            step_left = (x_vals[i] - x_vals[i - 1]) if i > 0 else (x_vals[1] - x_vals[0] if n > 1 else 1)
            step_right = (x_vals[j + 1] - x_vals[j]) if j + 1 < n else (x_vals[j] - x_vals[j - 1] if j > 0 else 1)
            left = x_vals[i] - step_left / 2
            right = x_vals[j] + step_right / 2
            ax.axvspan(left, right, color=color, alpha=alpha, zorder=0, linewidth=0)
            i = j + 1
        else:
            i += 1


def duennt_xticks_aus(member_index_werte, max_labels=14):
    """Reduziert die Anzahl angezeigter x-Achsenbeschriftungen (alle Datenpunkte bleiben erhalten),
    damit die Beschriftungen bei vielen Members pro Spannweite lesbar bleiben."""
    werte = list(member_index_werte)
    schritt = max(1, -(-len(werte) // max_labels))  # aufgerundete Ganzzahldivision
    return werte[::schritt]


# ==============================================================================
# ZEILEN-/DIAGRAMM-DEFINITION: je Nachweis eine Konfiguration
# ==============================================================================
DIAGRAMME = [
    {
        "suffix": "mu_max_vs_mEd",
        "spalten": ["mu_max", "mEd_p"],
        "labels": ["mu_max (MRd)", "mEd_p"],
        "ylabel": "Moment [kNm]",
        "bedingung": "med_groesser_als_mrd",
        "bedingung_label": "mEd_p > mu_max",
        "titel": "Biegewiderstand mu_max (MRd) vs. Bemessungsmoment mEd_p",
    },
    {
        "suffix": "vEd_vs_vu_p",
        "spalten": ["vu_p", "vEd"],
        "labels": ["vu_p (VRd)", "vEd"],
        "ylabel": "Querkraft [kN]",
        "bedingung": "ved_groesser_als_vrd",
        "bedingung_label": "vEd > vu_p",
        "titel": "Querkraftwiderstand vu_p (VRd) vs. Bemessungsquerkraft vEd",
    },
    {
        "suffix": "mkd_p_vs_mr_p",
        "spalten": ["mr_p", "mkd_p"],
        "labels": ["mr_p (Rissmoment)", "mkd_p"],
        "ylabel": "Moment [kNm]",
        "bedingung": "mkd_p_groesser_als_mr_p",
        "bedingung_label": "mkd_p > mr_p",
        "titel": "Rissmoment mr_p vs. Moment mkd_p (positive Biegung)",
    },
    {
        "suffix": "mkd_n_vs_mr_n",
        "spalten": ["mr_n", "mkd_n"],
        "labels": ["mr_n (Rissmoment)", "mkd_n"],
        "ylabel": "Moment [kNm]",
        "bedingung": "mkd_n_kleiner_als_mr_n",
        "bedingung_label": "mkd_n < mr_n",
        "titel": "Rissmoment mr_n vs. Moment mkd_n (negative Biegung)",
    },
    {
        "suffix": "w_install_adm_vs_w_install_ger",
        "spalten": ["w_install_adm", "w_install_ger"],
        "labels": ["w_install_adm (zulässig)", "w_install_ger (gerissen)"],
        "ylabel": "Durchbiegung w [m]",
        "bedingung": "w_install_ger_groesser_als_w_install_adm",
        "bedingung_label": "w_install_ger > w_install_adm",
        "titel": "Durchbiegung w_install_adm (zulässig) vs. w_install_ger (Durchbiegung gerissen)",
    },
    {
        "suffix": "w_use_adm_vs_w_use_ger",
        "spalten": ["w_use_adm", "w_use_ger"],
        "labels": ["w_use_adm (zulässig)", "w_use_ger (gerissen)"],
        "ylabel": "Durchbiegung w [m]",
        "bedingung": "w_use_ger_groesser_als_w_use_adm",
        "bedingung_label": "w_use_ger > w_use_adm",
        "titel": "Durchbiegung w_use_adm (zulässig) vs. w_use_ger (gerissen)",
    },
    {
        "suffix": "w_app_adm_vs_w_app_ger",
        "spalten": ["w_app_adm", "w_app_ger"],
        "labels": ["w_app_adm (zulässig)", "w_app_ger (gerissen)"],
        "ylabel": "Durchbiegung w [m]",
        "bedingung": "w_app_ger_groesser_als_w_app_adm",
        "bedingung_label": "w_app_ger > w_app_adm",
        "titel": "Durchbiegung w_app_adm (zulässig) vs. w_app_ger (gerissen)",
    },
]


# ==============================================================================
# PLOT: EINZELDIAGRAMM (SMALL MULTIPLES, EIN SUBPLOT JE SPANNWEITE)
# ==============================================================================
def plot_einzeldiagramm(df, spannweiten, diagramm_cfg, dataset_name, output_prefix):
    fig, axes = plt.subplots(1, len(spannweiten), figsize=(3.2 * len(spannweiten), 5), sharey=True)
    if len(spannweiten) == 1:
        axes = [axes]

    for ax, l in zip(axes, spannweiten):
        df_l = df[df["l_tot"] == l].sort_values("member_index")
        shade_condition_regions(ax, df_l["member_index"], df_l[diagramm_cfg["bedingung"]], FARBE_HINTERLEGUNG)
        for spalte, label, farbe in zip(diagramm_cfg["spalten"], diagramm_cfg["labels"], [FARBE_1, FARBE_2]):
            ax.plot(df_l["member_index"], df_l[spalte], marker="x", markersize=1,
                    linewidth=1, color=farbe, label=label)
        ax.set_title(f"l_tot = {l} m")
        ax.set_xlabel(f"Member Nr. (1–{int(df['member_index'].max())})")
        ax.set_xticks(duennt_xticks_aus(df_l["member_index"]))
        ax.tick_params(axis="x", labelrotation=90)
        ax.grid(True, linewidth=0.5, alpha=0.4)

    axes[0].set_ylabel(diagramm_cfg["ylabel"])

    handles, labels = axes[0].get_legend_handles_labels()
    handles.append(Patch(facecolor=FARBE_HINTERLEGUNG, alpha=0.6, label=diagramm_cfg["bedingung_label"]))
    labels.append(diagramm_cfg["bedingung_label"])
    fig.legend(handles, labels, loc="lower center", ncol=3, bbox_to_anchor=(0.5, -0.08))

    fig.suptitle(f"{diagramm_cfg['titel']} je Member, kategorisiert nach Spannweite\n({dataset_name})")
    fig.tight_layout(rect=[0, 0.06, 1, 1])

    for ext in ("png", "pdf"):
        pfad = f"{output_prefix}_{diagramm_cfg['suffix']}.{ext}"
        fig.savefig(pfad, dpi=300 if ext == "png" else None, bbox_inches="tight")
    plt.close(fig)
    print(f"Diagramm gespeichert: {output_prefix}_{diagramm_cfg['suffix']}.png / .pdf")


# ==============================================================================
# PLOT: KOMBINIERTE UEBERSICHT (3 ZEILEN x N SPALTEN)
# ==============================================================================
def plot_kombiniert(df, spannweiten, dataset_name, output_prefix):
    fig, axes = plt.subplots(len(DIAGRAMME), len(spannweiten),
                              figsize=(2.6 * len(spannweiten), 3.6 * len(DIAGRAMME)), sharex="col")
    if len(spannweiten) == 1:
        axes = axes.reshape(-1, 1)

    for row_idx, diagramm_cfg in enumerate(DIAGRAMME):
        for col_idx, l in enumerate(spannweiten):
            ax = axes[row_idx, col_idx]
            df_l = df[df["l_tot"] == l].sort_values("member_index")
            shade_condition_regions(ax, df_l["member_index"], df_l[diagramm_cfg["bedingung"]], FARBE_HINTERLEGUNG)
            for spalte, label, farbe in zip(diagramm_cfg["spalten"], diagramm_cfg["labels"], [FARBE_1, FARBE_2]):
                ax.plot(df_l["member_index"], df_l[spalte], marker="x", markersize=1,
                        linewidth=1, color=farbe, label=label)

            ax.grid(True, linewidth=0.5, alpha=0.4)
            ax.set_xticks(duennt_xticks_aus(df_l["member_index"], max_labels=8))
            ax.tick_params(axis="x", labelrotation=90)

            if row_idx == 0:
                ax.set_title(f"l_tot = {l} m")
            if row_idx == len(DIAGRAMME) - 1:
                ax.set_xlabel("Member Nr.")
            if col_idx == 0:
                ax.set_ylabel(diagramm_cfg["ylabel"])
                ax.legend(loc="best", framealpha=0.9, fontsize=9)

    hinterlegungs_labels = " / ".join(cfg["bedingung_label"] for cfg in DIAGRAMME)
    fig.legend(handles=[Patch(facecolor=FARBE_HINTERLEGUNG, alpha=0.6, label=hinterlegungs_labels)],
               loc="lower center", bbox_to_anchor=(0.5, -0.01))

    fig.suptitle(f"Uebersicht: Durchbiegung, Biegewiderstand und Querkraftwiderstand je Member, "
                 f"kategorisiert nach Spannweite\n({dataset_name})", fontsize=16)
    fig.tight_layout(rect=[0, 0.02, 1, 0.96])

    for ext in ("png", "pdf"):
        pfad = f"{output_prefix}_combined_overview.{ext}"
        fig.savefig(pfad, dpi=300 if ext == "png" else None, bbox_inches="tight")
    plt.close(fig)
    print(f"Diagramm gespeichert: {output_prefix}_combined_overview.png / .pdf")


# ==============================================================================
# MAIN
# ==============================================================================
def main():
    if len(sys.argv) >= 2:
        excel_file = sys.argv[1]
    else:
        skript_ordner = os.path.dirname(os.path.abspath(__file__))
        excel_file = os.path.join(skript_ordner, EXCEL_DATEI)
    sheet_name = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_SHEET_NAME

    output_dir = os.path.dirname(os.path.abspath(excel_file))
    dataset_name = os.path.splitext(os.path.basename(excel_file))[0]
    output_prefix = os.path.join(output_dir, dataset_name)

    df = lade_daten(excel_file, sheet_name)
    spannweiten = sorted(df["l_tot"].unique())

    for diagramm_cfg in DIAGRAMME:
        plot_einzeldiagramm(df, spannweiten, diagramm_cfg, dataset_name, output_prefix)
    plot_kombiniert(df, spannweiten, dataset_name, output_prefix)


if __name__ == "__main__":
    main()
