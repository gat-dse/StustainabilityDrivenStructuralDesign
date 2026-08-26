"""
Liest die Rohdaten-Excel "Members_rc_rib_simple_massiv_bvar.xlsx" (Sheet "Members_Comparison") ein.
Gleicher Rohaufbau wie die anderen rc_rib-Dateien (Spalte 'key' + 'level', je eine Spalte pro Member).

Es werden nur die fuer die drei Nachweis-Diagramme benoetigten Groessen herausgezogen, transponiert
(eine Zeile je Member) und als SQLite-Datenbank gespeichert:
- w_use / w_use_ger (Durchbiegung, gerissen/ungerissen)
- mu_max(MRd) / mEd_p (Biegewiderstand vs. Bemessungsmoment, positive Biegung)
- vu_PB_p(VRd) / vEd (Querkraftwiderstand vs. Bemessungsquerkraft)
"""
import sqlite3

import openpyxl
import pandas as pd

excel_file = "Members_rc_rib_simple_massiv_bvar.xlsx"
sheet_name = "Members_Comparison"
db_file = "Members_rc_rib_simple_massiv_bvar.db"
table_name = "members"

# ==============================================================================
# EINLESEN & BENOETIGTE ZEILEN FINDEN (erstes Vorkommen je (key, level))
# ==============================================================================
wb = openpyxl.load_workbook(excel_file, data_only=True)
ws = wb[sheet_name]

# (key, level)
benoetigte_keys = [
    ("l_tot", 1),
    ("w_use", 0), ("w_use_ger", 0),
    ("mu_max(MRd)", 1), ("mEd_p", 0),
    ("vu_PB_p(VRd)", 1), ("vEd", 0),
]
zeilen_index = {}
for r in range(1, ws.max_row + 1):
    key = ws.cell(row=r, column=1).value
    level = ws.cell(row=r, column=2).value
    if (key, level) in benoetigte_keys and key not in zeilen_index:
        zeilen_index[key] = r

member_ids = [ws.cell(row=1, column=c).value for c in range(3, ws.max_column + 1)]

daten = {"Member_ID": member_ids}
for key, r in zeilen_index.items():
    daten[key] = [ws.cell(row=r, column=c).value for c in range(3, ws.max_column + 1)]

df = pd.DataFrame(daten)

# ==============================================================================
# SPEICHERN ALS SQLITE-DATENBANK
# ==============================================================================
con = sqlite3.connect(db_file)
df.to_sql(table_name, con, if_exists="replace", index=False)
con.close()

print(f"Datenbank erfolgreich erstellt: {db_file}")
print(f"Tabelle '{table_name}': {df.shape[0]} Zeilen (Members) x {df.shape[1]} Spalten")
