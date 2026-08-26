"""
Liest die Rohdaten-Excel "Members_rc_rib_simple_massiv_bvar_Anm.xlsx" (Sheet "Members_Comparison") ein.
Gleicher Aufbau wie bei Members_rc_rec_simple_massiv (Spalte 'key' + 'level', je eine Spalte
pro Member), aber fuer Rippenplatten (rc_rib) mit zusaetzlichen, hier nicht benoetigten Zeilen.

Es werden nur die fuer die Diagramme benoetigten Groessen herausgezogen, transponiert
(eine Zeile je Member) und als SQLite-Datenbank gespeichert - mit denselben Spaltennamen
wie Members_rc_rec_simple_massiv.db, damit dieselben Plot-Skripte wiederverwendet werden koennen.

Besonderheiten dieser Rohdatei:
- 'Faktor_ger' entspricht 'fwger(phi=2)_calc', da phi fuer alle Members = 2 ist.
- 'h/d_calc' ist nicht direkt enthalten und wird aus den Rohzeilen 'h' und 'd' berechnet.
- 'co2' kommt zweimal vor (level 1: nur Querschnitt, level 0: gesamter Member inkl.
  Bodenaufbau, in kgCO2eq/m2). Hier wird gezielt level 0 verwendet.
- 'prod_id' / 'GWP' (level 2) kommen je zweimal vor: 1. = Beton, 2. = Bewehrung.
"""
import sqlite3

import openpyxl
import pandas as pd

excel_file = "Members_rc_rib_simple_massiv_bvar_Anm.xlsx"
sheet_name = "Members_Comparison"
db_file = "Members_rc_rib_simple_massiv.db"
table_name = "members"

# ==============================================================================
# EINLESEN & BENOETIGTE ZEILEN FINDEN (erstes Vorkommen je (key, level))
# ==============================================================================
wb = openpyxl.load_workbook(excel_file, data_only=True)
ws = wb[sheet_name]

# (key, level) - level wird benoetigt, da einige Keys (z.B. 'co2') mehrfach vorkommen
benoetigte_keys = [
    ("h", 1), ("d", 1), ("roh", 1), ("rohs", 1), ("Faktor_ger", 1),
    ("l_tot", 1), ("w_use", 0), ("w_use_ger", 0), ("co2", 0),
]
zeilen_index = {}
for r in range(1, ws.max_row + 1):
    key = ws.cell(row=r, column=1).value
    level = ws.cell(row=r, column=2).value
    if (key, level) in benoetigte_keys and key not in zeilen_index:
        zeilen_index[key] = r


# 'prod_id' / 'GWP' (level 2) kommen je zweimal vor: 1. = Beton, 2. = Bewehrung (siehe Docstring)
def zwei_zeilen(key):
    zeilen = [r for r in range(1, ws.max_row + 1)
              if ws.cell(row=r, column=1).value == key and ws.cell(row=r, column=2).value == 2]
    assert len(zeilen) == 2, f"Erwartete genau 2 '{key}'-Zeilen (Beton, Bewehrung)"
    return zeilen


zeile_prod_id_concrete, zeile_prod_id_rebar = zwei_zeilen("prod_id")
zeile_gwp_concrete, zeile_gwp_rebar = zwei_zeilen("GWP")

member_ids = [ws.cell(row=1, column=c).value for c in range(3, ws.max_column + 1)]

daten = {"Member_ID": member_ids}
for key, r in zeilen_index.items():
    daten[key] = [ws.cell(row=r, column=c).value for c in range(3, ws.max_column + 1)]
for spalte, r in [("prod_id_concrete", zeile_prod_id_concrete), ("prod_id_rebar", zeile_prod_id_rebar),
                   ("gwp_concrete", zeile_gwp_concrete), ("gwp_rebar", zeile_gwp_rebar)]:
    daten[spalte] = [ws.cell(row=r, column=c).value for c in range(3, ws.max_column + 1)]

df = pd.DataFrame(daten)

# ==============================================================================
# ABLEITUNG DER FEHLENDEN GROESSEN
# ==============================================================================
df = df.rename(columns={"Faktor_ger": "fwger(phi=2)_calc"})
df["h/d_calc"] = df["h"] / df["d"]

# ==============================================================================
# SPEICHERN ALS SQLITE-DATENBANK
# ==============================================================================
con = sqlite3.connect(db_file)
df.to_sql(table_name, con, if_exists="replace", index=False)
con.close()

print(f"Datenbank erfolgreich erstellt: {db_file}")
print(f"Tabelle '{table_name}': {df.shape[0]} Zeilen (Members) x {df.shape[1]} Spalten")
